# -*- coding: utf-8 -*-
import re
import unicodedata
from collections import Counter, defaultdict
from itertools import combinations
from typing import Dict, List, Tuple, Set, Optional

from janome.tokenizer import Tokenizer
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib
import os
import math
from pathlib import Path

# Excel 出力用
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================================
# 設定（必要なら変更してください）
# =========================================================
DEBUG = False
COOCCURRENCE_MIN_FREQ = 5
COOCCURRENCE_WINDOW = 5
DRAW_TOP_EDGES = 40
MAX_NEIGHBORS_FOR_SET = 20
TOP_PROBLEM_SET_PER_TEGUCHI = 5
FONT_FAMILY = "Yu Gothic"

TOP_PAIR_CANDIDATES_PER_PERIOD = 10
TOP_OVERALL_PAIRS = 40
PLOT_TOP_PAIRS = 10
TOP_COOC_PER_PERIOD = 10

PER_PERIOD_KEEP = 1
FILL_FROM_OVERALL = 4
NORMALIZE_TIMESERIES = True

TEGUCHI_COLOR = "#d97474"
PROBLEM_COLOR = "#7297b4"

# 出力フォルダ
OUTPUT_DIR = Path("results_cooccurrence")
NETWORKS_DIR = OUTPUT_DIR / "networks"
CHARTS_DIR = OUTPUT_DIR / "charts"
TABLES_DIR = OUTPUT_DIR / "tables"

# 折れ線プロット関連の制約
MAX_TIMESERIES = 10            # 折れ線で表示する"系列"の最大数（ユーザー要望）
MIN_ABS_CHANGE = 2             # 絶対差の閾値（max-min がこれ以上ない系列は除外）
MIN_REL_CHANGE = 0.20          # 相対変化の閾値 ((max-min)/max) がこれ以上ない系列は除外（max>0時）
USE_REL_OR_ABS = True          # True: 両方を満たすものだけ除外, False: 絶対差だけで判定

# ネットワーク図の保存数
MAX_NETWORK_PLOTS = 6         # 保存する共起ネットワーク図の最大数

# =========================================================
# 前処理（サイト -> ネット 統一）
# =========================================================
class TextPreprocessor:
    def __init__(self, stopwords_file: str = "stopwords.txt"):
        self.tokenizer = Tokenizer()
        self.stop_words = set()
        self.stopwords_file = stopwords_file
        self._load_stopwords()

    def _load_stopwords(self):
        if os.path.exists(self.stopwords_file):
            with open(self.stopwords_file, encoding="utf-8") as f:
                self.stop_words |= set(f.read().splitlines())
        else:
            open(self.stopwords_file, "w", encoding="utf-8").close()

    def clean_text(self, text: str) -> str:
        text = re.sub(r'https?://\S+', '', str(text))
        text = re.sub(r'\d{2,4}年\d{1,2}月\d{1,2}日', '', text)
        text = re.sub(r'\d{2,4}[-/]\d{1,2}[-/]\d{1,2}', '', text)
        text = re.sub(r'\d{2,4}-\d{2,4}-\d{2,4}', '', text)
        return text

    def tokenize(self, text: str) -> List[str]:
        text = self.clean_text(text)
        tokens: List[str] = []
        for token in self.tokenizer.tokenize(text):
            pos = token.part_of_speech
            base = token.base_form
            if pos.startswith("名詞") and not pos.startswith(("名詞,数", "名詞,代名詞")) and len(base) > 1:
                if base not in self.stop_words and not re.match(r'^[0-9０-９]+$', base):
                    if base == "サイト":
                        base = "ネット"
                    if base == "携帯":
                        base = "電話"
                    tokens.append(base)
        return tokens

# =========================================================
# ModalityClassifier（TEGUCHI_MAPのみ厳密に使用）
# =========================================================
class ModalityClassifier:
    def __init__(self, teguchi_map: Dict[str, List[str]]):
        self.teguchi_map = teguchi_map
        self.word2teguchi: Dict[str, str] = {}
        for label, keywords in teguchi_map.items():
            for w in keywords:
                self.word2teguchi[w] = label

    def is_teguchi(self, word: str) -> bool:
        return word in self.word2teguchi

    def get_label(self, word: str) -> Optional[str]:
        return self.word2teguchi.get(word, None)

# =========================================================
# CSV読み込み / 期間分割 / トークン化 / 共起計算（元ロジック維持）
# =========================================================
def load_csv(file_path: str = "R5 1.csv") -> pd.DataFrame:
    df = pd.read_csv(file_path, usecols=[0,6], skiprows=1, header=None, encoding="cp932")
    df.columns = ["date","text"]
    df["date"] = pd.to_datetime(df["date"], format="%Y年%m月%d日", errors="coerce")
    return df.dropna(subset=["date"]) 

def split_periods(df: pd.DataFrame):
    start_date = df["date"].min()
    end_date = df["date"].max()
    if pd.isna(start_date) or pd.isna(end_date):
        return []
    current = start_date.replace(day=1)
    periods = []
    while current <= end_date:
        period_start = current
        period_end = min(period_start + pd.DateOffset(months=2) - pd.DateOffset(days=1), end_date)
        chunk = df[(df["date"] >= period_start) & (df["date"] <= period_end)]
        periods.append((period_start, period_end, chunk))
        current = (period_start + pd.DateOffset(months=2)).replace(day=1)
    return periods

def tokenize_corpus(docs: List[str], preprocessor: TextPreprocessor):
    processed_docs = [preprocessor.tokenize(d) for d in docs if isinstance(d, str)]
    all_tokens = [t for doc in processed_docs for t in doc]
    freq = Counter(all_tokens)
    tokenized_docs = [[t for t in doc if 3 <= freq[t] <= 500] for doc in processed_docs]
    return tokenized_docs, freq

def build_cooccurrence_counter(tokenized_docs: List[List[str]], window_size=None):
    cooccurrence = Counter()
    for tokens in tokenized_docs:
        if not tokens: continue
        if window_size:
            for i in range(len(tokens)):
                for j in range(i+1, min(i+window_size+1, len(tokens))):
                    if tokens[i] != tokens[j]:
                        pair = tuple(sorted((tokens[i], tokens[j])))
                        cooccurrence[pair] += 1
        else:
            uniq = list(set(tokens))
            for w1, w2 in combinations(uniq, 2):
                pair = tuple(sorted((w1, w2)))
                cooccurrence[pair] += 1
    return cooccurrence

def build_cooccurrence_network_from_counter(counter: Counter, min_freq=5):
    filtered = {pair: c for pair, c in counter.items() if c >= min_freq}
    G = nx.Graph()
    for (w1, w2), c in filtered.items():
        G.add_edge(w1, w2, weight=c)
    return G

# =========================================================
# 表作成（改良：表示用とペアキー用の DataFrame を返す）
# =========================================================
def build_cooccurrence_top_table(period_labels, period_cooccurrence_counters, top_n=TOP_COOC_PER_PERIOD):
    period_top_pairs_display = []
    period_top_pairs_key = []
    for c in period_cooccurrence_counters:
        top = c.most_common(top_n)
        # 表示・キーは半角 '|' で統一
        formatted = [f"{p[0][0]}|{p[0][1]} ({p[1]})" for p in top]
        keys = [f"{p[0][0]}|{p[0][1]}" for p in top]
        formatted += [""] * max(0, top_n - len(formatted))
        keys += [""] * max(0, top_n - len(keys))
        period_top_pairs_display.append(formatted[:top_n])
        period_top_pairs_key.append(keys[:top_n])
    index = [f"{i+1}" for i in range(top_n)]
    df_display = pd.DataFrame({label: period_top_pairs_display[i] for i, label in enumerate(period_labels)}, index=index)
    df_keys = pd.DataFrame({label: period_top_pairs_key[i] for i, label in enumerate(period_labels)}, index=index)
    return df_display, df_keys

# =========================================================
# ユーティリティ: キー正規化と色生成
# =========================================================
def _normalize_key(s: Optional[str]) -> str:
    """テーブルキー文字列を比較用に正規化して返す（NFKC + trim）。"""
    if s is None:
        return ""
    s2 = unicodedata.normalize("NFKC", str(s))
    s2 = s2.replace("｜", "|").replace("　", " ").strip()
    s2 = re.sub(r"\s+", " ", s2)
    return s2

def _generate_distinct_palette(n: int, alpha: float = 0.25):
    import matplotlib.colors as mcolors
    if n <= 0:
        return []
    colors = []
    for i in range(n):
        h = float(i) / float(n)
        s = 0.65
        v = 0.92
        rgb = mcolors.hsv_to_rgb((h, s, v))
        rgba = (float(rgb[0]), float(rgb[1]), float(rgb[2]), float(alpha))
        colors.append(rgba)
    return colors

def _extract_pair_from_display_cell(display_cell: Optional[str]) -> str:
    if display_cell is None:
        return ""
    s = str(display_cell).strip()
    if not s:
        return ""
    m = re.match(r'^(.*?)\s*\(\s*\d+\s*\)\s*$', s)
    base = m.group(1) if m else s
    base = base.replace("｜", "|")
    return _normalize_key(base)

# =========================================================
# 表表示（ユニークのみハイライト可能にしたバージョン）
# =========================================================
def show_cooccurrence_table(df_cooc_display: pd.DataFrame,
                            df_cooc_keys: pd.DataFrame = None,
                            alpha: float = 0.25,
                            save_path: Optional[Path] = None,
                            highlight_unique_only: bool = True):
    print("\n=== 期間ごとの上位共起ペア（順位 1 → {}） ===".format(df_cooc_display.shape[0]))
    print(df_cooc_display.to_string())

    # 期間ごとにどのペアが出現しているかを集計（df_cooc_keys を優先して利用）
    pair_period_counts = {}
    if df_cooc_keys is not None:
        cols = df_cooc_keys.columns
        for col in cols:
            for raw in df_cooc_keys[col].values:
                if raw is None:
                    continue
                s = str(raw).strip()
                if not s:
                    continue
                nk = _normalize_key(s.replace("｜", "|"))
                if nk:
                    pair_period_counts[nk] = pair_period_counts.get(nk, 0) + 1
    # フォールバック（もし df_cooc_keys が完全でない場合）
    for raw in df_cooc_display.values.flatten():
        if raw is None:
            continue
        s = str(raw).strip()
        if not s:
            continue
        nk = _extract_pair_from_display_cell(s)
        if nk and nk not in pair_period_counts:
            pair_period_counts[nk] = pair_period_counts.get(nk, 0) + 1

    unique_pairs = [p for p, cnt in pair_period_counts.items() if cnt == 1]

    n = max(1, len(unique_pairs))
    palette = _generate_distinct_palette(n, alpha=alpha)
    color_map = {pair: palette[i] for i, pair in enumerate(unique_pairs)}

    fig, ax = plt.subplots(figsize=(max(8, df_cooc_display.shape[1]*1.6), 1 + df_cooc_display.shape[0]*0.6))
    ax.axis('off')
    table = ax.table(cellText=df_cooc_display.values,
                     colLabels=df_cooc_display.columns,
                     rowLabels=df_cooc_display.index,
                     cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 1.2)

    cell_map = {}
    for key, cell in table.get_celld().items():
        txt = cell.get_text().get_text()
        if not txt:
            continue
        base = _extract_pair_from_display_cell(str(txt).strip())
        if base:
            cell_map.setdefault(base, []).append(key)

    used_keys = set()
    rows = df_cooc_display.shape[0]
    cols = df_cooc_display.shape[1]
    for r in range(rows):
        for c in range(cols):
            try:
                key_val = ""
                if df_cooc_keys is not None:
                    raw_key = df_cooc_keys.iloc[r, c]
                    if raw_key is not None and str(raw_key).strip():
                        key_val = _normalize_key(str(raw_key).replace("｜", "|"))

                if not key_val:
                    raw_display = df_cooc_display.iloc[r, c]
                    key_val = _extract_pair_from_display_cell(raw_display)

                if not key_val:
                    continue

                if highlight_unique_only and pair_period_counts.get(key_val, 0) != 1:
                    continue

                rgba = color_map.get(key_val)
                if rgba is None:
                    continue

                candidates = cell_map.get(key_val, [])
                chosen = None
                for k in candidates:
                    if k not in used_keys:
                        chosen = k
                        break
                if chosen is None and candidates:
                    chosen = candidates[0]
                if chosen is not None:
                    used_keys.add(chosen)
                    cell = table[chosen]
                    cell.set_facecolor(rgba)
                    cell.set_edgecolor("black")
                    cell.get_text().set_color("black")
            except Exception as e:
                if DEBUG:
                    print(f"[error] r={r} c={c} -> {e}")

    plt.title(f"各期間の上位 {df_cooc_display.shape[0]} 共起ペアランキング（ユニークのみハイライト）")
    plt.tight_layout()
    if save_path is not None:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(save_path, dpi=200, bbox_inches='tight')
        print(f"共起表画像を保存しました: {save_path}")
    plt.show()
    plt.close(fig)

# =========================================================
# Excel 出力（ユニークのみハイライト）
# =========================================================
def _blend_hex_with_white(hexcol: str, alpha: float) -> str:
    if not hexcol or not hexcol.startswith("#") or len(hexcol) != 7:
        return "#FFFFFF"
    r = int(hexcol[1:3], 16)
    g = int(hexcol[3:5], 16)
    b = int(hexcol[5:7], 16)
    r2 = int((1.0 - alpha) * 255 + alpha * r)
    g2 = int((1.0 - alpha) * 255 + alpha * g)
    b2 = int((1.0 - alpha) * 255 + alpha * b)
    return "{:02X}{:02X}{:02X}".format(r2, g2, b2)

def export_cooccurrence_table_to_excel(df_keys: pd.DataFrame,
                                       df_display: pd.DataFrame,
                                       filename: str = "cooccurrence_top.xlsx",
                                       alpha: float = 0.25,
                                       size_multiplier: float = 2.0,
                                       highlight_unique_only: bool = True):
    # 期間ごとにどのペアが出現しているかを集計
    pair_period_counts = {}
    if df_keys is not None:
        cols = df_keys.columns
        for col in cols:
            for raw in df_keys[col].values:
                if raw is None:
                    continue
                s = str(raw).strip()
                if not s:
                    continue
                nk = _normalize_key(s)
                if nk:
                    pair_period_counts[nk] = pair_period_counts.get(nk, 0) + 1

    unique_pairs = [p for p, cnt in pair_period_counts.items() if cnt == 1]

    n = max(1, len(unique_pairs))
    import matplotlib.colors as mcolors
    palette = _generate_distinct_palette(n, alpha=alpha)
    base_color_map = {}
    for i, pair in enumerate(unique_pairs):
        rgb = palette[i][:3]
        hexcol = mcolors.to_hex(rgb)
        base_color_map[pair] = hexcol

    blended_map = {pair: ('#' + _blend_hex_with_white(base_color_map[pair], alpha)) for pair in base_color_map}

    wb = Workbook()
    ws = wb.active
    ws.title = "共起上位"

    cols = list(df_display.columns)
    header = ["順位\\期間"] + cols
    ws.append(header)

    for row_idx in range(df_display.shape[0]):
        row = [df_display.index[row_idx]]
        for col_idx, col in enumerate(cols):
            display_val = df_display.iloc[row_idx, col_idx]
            row.append(display_val)
        ws.append(row)

    base_col_width = 20
    base_header_height = 18
    base_row_height = 15
    final_col_width = base_col_width * max(0.5, size_multiplier)
    final_header_height = base_header_height * max(0.5, size_multiplier)
    final_row_height = base_row_height * max(0.5, size_multiplier)

    for ci in range(1, len(header) + 1):
        col_letter = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[col_letter].width = final_col_width

    ws.row_dimensions[1].height = final_header_height
    for r in range(2, 2 + df_display.shape[0]):
        ws.row_dimensions[r].height = final_row_height

    rows = df_display.shape[0]
    cols_count = len(cols)
    for r in range(rows):
        for c in range(cols_count):
            raw_key = df_keys.iloc[r, c]
            key_val = _normalize_key("" if raw_key is None else str(raw_key))
            if not key_val:
                continue
            if highlight_unique_only and pair_period_counts.get(key_val, 0) != 1:
                continue
            blended_hex = blended_map.get(key_val, "#FFFFFF")
            argb = "FF" + blended_hex[1:].upper()
            fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
            cell = ws.cell(row=2 + r, column=2 + c)
            cell.fill = fill
            cell.font = Font(color="FF000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    legend_ws = wb.create_sheet(title="凡例")
    legend_ws.append(["色（合成）", "ペア"])
    legend_ws.column_dimensions[legend_ws.cell(row=1, column=1).column_letter].width = final_col_width
    legend_ws.column_dimensions[legend_ws.cell(row=1, column=2).column_letter].width = final_col_width * 1.5
    for pair in unique_pairs:
        blended_hex = blended_map.get(pair, "#FFFFFF")
        argb = "FF" + blended_hex[1:].upper()
        fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
        legend_ws.append([pair, pair])
        r = legend_ws.max_row
        legend_ws.cell(row=r, column=1).fill = fill
        legend_ws.cell(row=r, column=2).value = pair
        legend_ws.row_dimensions[r].height = final_row_height

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    wb.save(path)
    print(f"Excel ファイルを保存しました: {path}  (alpha={alpha}, size_multiplier={size_multiplier})")
    return path

# =========================================================
# 問題セット抽出（変更点：3共起のみを参照するように変更）
# =========================================================
def _score_set(G: nx.Graph, teguchi: Optional[str], problems: Tuple[str, ...]) -> float:
    """
    変更点:
    - 前と同じ（最小値を返すロジック）。
    - ただし、呼び出し側を 3 共起用に限定して使います（関数自体は汎用）。
    """
    weights = []

    # 手口と各問題間のエッジ（存在しないならセットとして同時出現は成立しない）
    if teguchi:
        for p in problems:
            if G.has_edge(teguchi, p):
                weights.append(G[teguchi][p]["weight"])
            else:
                return 0.0

    # 問題間のペアエッジ（全て存在する必要がある）
    for a, b in combinations(problems, 2):
        if G.has_edge(a, b):
            weights.append(G[a][b]["weight"])
        else:
            return 0.0

    if not weights:
        # 必要エッジが全く見つからない場合のフォールバック
        if teguchi is None and len(problems) == 1:
            p = problems[0]
            # フォールバック: その語の最大隣接エッジ重み（最も強い共起）
            maxw = 0
            if p in G:
                for nbr in G.neighbors(p):
                    w = G[p][nbr]["weight"]
                    if w > maxw:
                        maxw = w
            return float(maxw)
        return 0.0

    # 保守的下界として最小値を返す
    return float(min(weights))

def extract_problem_sets_for_teguchi(
    G: nx.Graph,
    teguchi_node: str,
    classifier: ModalityClassifier,
    max_neighbors: int = MAX_NEIGHBORS_FOR_SET
) -> Dict[Tuple[str, ...], float]:
    """
    変更点（重要）:
    - ここでは「手口ノード + 2つの問題語（= 合計3語での共起）」のみを抽出します。
      つまり problems 引数には長さ2のタプル (p1, p2) を与え _score_set を呼びます。
    - 単独語や問題1語のみのセットは作りません（要望どおり3共起のみ参照）。
    """
    if teguchi_node not in G:
        return {}
    neighbors = [n for n in G.neighbors(teguchi_node) if not classifier.is_teguchi(n)]
    neighbors.sort(key=lambda n: G[teguchi_node][n]["weight"], reverse=True)
    problems = neighbors[:max_neighbors]
    scores: Dict[Tuple[str, ...], float] = {}

    # ペア (p1, p2) を作って、teguchi + (p1,p2) の 3 共起のみ評価
    for p1, p2 in combinations(problems, 2):
        if classifier.is_teguchi(p1) or classifier.is_teguchi(p2):
            continue
        tpl = tuple(sorted((p1, p2)))
        s = _score_set(G, teguchi_node, tpl)
        if s > 0:
            scores[tpl] = s
    return scores

def summarize_top_problem_sets_with_others(G: nx.Graph, classifier: ModalityClassifier, top_k: int = 5):
    """
    変更点（重要）:
    - 各手口については extract_problem_sets_for_teguchi を用い、手口 + 2語（=3共起）だけを表示。
    - 'その他' は問題語3つ (triplet) のみを評価して上位を表示する（これも 3 共起に限定）。
    """
    summaries = {}
    used_nodes: Set[str] = set()
    for node in G.nodes:
        if classifier.is_teguchi(node):
            teguchi_label = classifier.get_label(node) or node
            scores = extract_problem_sets_for_teguchi(G, node, classifier)
            top_items = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:top_k]
            summaries[teguchi_label] = top_items
            for tpl, _ in top_items:
                used_nodes.update(tpl)
    other_nodes = [n for n in G.nodes if not classifier.is_teguchi(n) and n not in used_nodes]
    other_scores: Dict[Tuple[str, ...], float] = {}
    neighbors = other_nodes[:MAX_NEIGHBORS_FOR_SET]
    # ここで triplet（問題語3つ）の組み合わせのみ評価する
    for combo in combinations(neighbors, 3):
        # 排除: 中に手口語が入っていたらスキップ
        if any(classifier.is_teguchi(x) for x in combo):
            continue
        tpl = tuple(sorted(combo))
        s = _score_set(G, None, tpl)
        if s > 0:
            other_scores[tpl] = s
    top_other_items = sorted(other_scores.items(), key=lambda x: x[1], reverse=True)[:top_k]
    if top_other_items:
        summaries["その他"] = top_other_items
    return summaries

# =========================================================
# 折れ線の候補選出（手口語を含むペアは完全に除外）
# =========================================================
def select_candidate_pairs_improved(period_cooccurrence_counters: List[Counter],
                                    period_labels: List[str],
                                    classifier: ModalityClassifier,
                                    per_period_keep: int = PER_PERIOD_KEEP,
                                    fill_from_overall: int = FILL_FROM_OVERALL,
                                    exclude_any_teguchi_in_pair: bool = True) -> List[Tuple[str, str]]:
    selected = []
    for c in period_cooccurrence_counters:
        top = [pair for pair, _ in c.most_common()]
        kept = 0
        for pair in top:
            w1, w2 = pair
            if exclude_any_teguchi_in_pair and (classifier.is_teguchi(w1) or classifier.is_teguchi(w2)):
                continue
            if pair not in selected:
                selected.append(pair)
                kept += 1
            if kept >= per_period_keep:
                break
    total = Counter()
    for c in period_cooccurrence_counters:
        total.update(c)
    for pair, _ in total.most_common():
        if len(selected) >= per_period_keep * len(period_cooccurrence_counters) + fill_from_overall:
            break
        w1, w2 = pair
        if exclude_any_teguchi_in_pair and (classifier.is_teguchi(w1) or classifier.is_teguchi(w2)):
            continue
        if pair in selected:
            continue
        selected.append(pair)
    uniq = []
    seen = set()
    for p in selected:
        if p not in seen:
            uniq.append(p)
            seen.add(p)
    if DEBUG:
        print("候補ペア（最終）:", uniq[:50])
    return uniq

# =========================================================
# 時系列データ作成（正規化オプション）
# 変更: 分母を「期間の行数（相談件数）」にするように period_doc_counts を使う
# =========================================================
def build_pair_timeseries_with_norm(period_labels: List[str],
                                    period_cooccurrence_counters: List[Counter],
                                    candidate_pairs: List[Tuple[str, str]],
                                    period_doc_counts: Optional[List[int]] = None,
                                    normalize: bool = NORMALIZE_TIMESERIES) -> pd.DataFrame:
    rows = []
    # 変更点: totals を各期間の行数（相談件数）にする
    if period_doc_counts and len(period_doc_counts) == len(period_cooccurrence_counters):
        totals = period_doc_counts
    else:
        # フォールバック: cooccurrence の合計を使う（ただし今回の要望では period_doc_counts を渡してください）
        totals = [sum(c.values()) if c else 0 for c in period_cooccurrence_counters]

    for idx, c in enumerate(period_cooccurrence_counters):
        row = {}
        for pair in candidate_pairs:
            val = c.get(pair, 0)
            if normalize and totals[idx] > 0:
                row["|".join(pair)] = val / totals[idx]
            else:
                row["|".join(pair)] = val
        rows.append(row)
    df = pd.DataFrame(rows, index=period_labels)
    return df.fillna(0)

# =========================================================
# 折れ線グラフ（フィルタリング機能と PNG 保存対応）
# =========================================================
def filter_timeseries_by_change(df_pairs: pd.DataFrame,
                                 min_abs: float = MIN_ABS_CHANGE,
                                 min_rel: float = MIN_REL_CHANGE,
                                 use_rel_or_abs: bool = USE_REL_OR_ABS) -> pd.DataFrame:
    if df_pairs.empty:
        return df_pairs
    keep_cols = []
    for col in df_pairs.columns:
        vals = df_pairs[col].values
        mx = vals.max()
        mn = vals.min()
        diff = mx - mn
        rel = (diff / mx) if mx > 0 else 0.0
        if use_rel_or_abs:
            if diff >= min_abs or rel >= min_rel:
                keep_cols.append(col)
        else:
            if diff >= min_abs:
                keep_cols.append(col)
    return df_pairs[keep_cols]

def plot_pair_timeseries_improved(df_pairs: pd.DataFrame, top_n: Optional[int] = PLOT_TOP_PAIRS, normalize: bool = False, save_path: Optional[Path] = None):
    if df_pairs.empty:
        print("共起時系列データが空です。")
        return

    df_filtered = filter_timeseries_by_change(df_pairs)
    if df_filtered.empty:
        print("（フィルタにより表示対象がなくなりました — 閾値を下げるか設定を変更してください）")
        return

    total = df_filtered.sum(axis=0).sort_values(ascending=False)
    limit = min(MAX_TIMESERIES, top_n if top_n and top_n > 0 else len(total))
    cols = total.index.tolist()[:limit]

    print("\n--- 折れ線でプロットする共起ペア ---")
    for c in cols:
        print("  ", c)

    fig, ax = plt.subplots(figsize=(12, 6))
    x = list(range(len(df_filtered.index)))
    for col in cols:
        ax.plot(x, df_filtered[col].values, marker='o', label=col)
    ax.set_xticks(x)
    ax.set_xticklabels(df_filtered.index, rotation=45)
    ax.set_xlabel("期間")
    ax.set_ylabel("共起頻度" + ("（割合）" if normalize else ""))
    ax.set_title("候補共起ペアの期間推移")
    ax.legend(loc='upper left', bbox_to_anchor=(1.02, 1))
    plt.tight_layout()
    if save_path is not None:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(save_path, dpi=200, bbox_inches='tight')
        print(f"時系列図を保存しました: {save_path}")
    plt.show()
    plt.close(fig)

# =========================================================
# 追加: 積み上げ棒グラフ表示関数（縦軸を割合(0..1)で表示、かつ「その他」を灰色で追加）
# 変更点:
# - normalize=True のとき、各期間の分母は「その期間の行数（相談件数）」を使って build_pair_timeseries_with_norm が正規化済みの値を返す想定
# - 上位表示の合計が総数に満たない部分を 'その他' として灰色で追加
# - 色の割当順は元コードの要望に合わせて逆順にしている
# =========================================================
def plot_pair_stacked_bars(df_pairs: pd.DataFrame,
                           top_n: Optional[int] = PLOT_TOP_PAIRS,
                           normalize: bool = False,
                           save_path: Optional[Path] = None):
    """
    積み上げ棒グラフ（色の割り当て順を確実に逆にする & 凡例順を積み上げ順に合わせる）
    normalize=True のときは各期間を 0..1 の割合に正規化して表示（パーセンテージ表示ではない）。
    """
    if df_pairs.empty:
        print("共起時系列データが空です。")
        return

    df_filtered = filter_timeseries_by_change(df_pairs)
    if df_filtered.empty:
        print("（フィルタにより表示対象がなくなりました — 閾値を下げるか設定を変更してください）")
        return

    total = df_filtered.sum(axis=0).sort_values(ascending=False)
    limit = min(MAX_TIMESERIES, top_n if top_n and top_n > 0 else len(total))
    cols = total.index.tolist()[:limit]

    if not cols:
        print("表示する系列がありません。")
        return

    df_plot = df_filtered[cols].copy()
    if normalize:
        # df_plot の各行は build_pair_timeseries_with_norm で
        # 「期間の行数」を分母にして正規化済みのはず（本関数では再正規化しない）
        displayed_sum = df_plot.sum(axis=1)
        others = (1.0 - displayed_sum).clip(lower=0.0)
        df_plot = pd.concat([others.rename("その他"), df_plot], axis=1)
        new_cols = df_plot.columns.tolist()
    else:
        new_cols = cols  # 変更なし（'その他' は付けない）

    # カラーパレット（元順）→ 明示的に逆順にして列にマップ
    colors = _generate_distinct_palette(len(cols), alpha=0.9)
    colors_rev = list(reversed(colors))
    color_map = {}
    if normalize:
        # 'その他' は灰色（目立ち過ぎないように rgba で設定）
        color_map["その他"] = (0.7, 0.7, 0.7, 1.0)
        for i, col in enumerate(cols):
            color_map[col] = colors_rev[i]
    else:
        for i, col in enumerate(cols):
            color_map[col] = colors_rev[i]

    x = list(range(len(df_plot.index)))
    fig, ax = plt.subplots(figsize=(max(10, df_plot.shape[1] * 1.2), 6))
    border_width = 0.6

    bottoms = [0.0] * len(df_plot.index)
    # 描画順は new_cols の順（'その他' を先に入れたので、それが下段になる）
    for col in new_cols:
        values = df_plot[col].values
        ax.bar(x, values, bottom=bottoms, label=col, width=0.6, alpha=1.0,
               color=color_map.get(col, None), edgecolor='black', linewidth=border_width)
        bottoms = [b + v for b, v in zip(bottoms, values)]

    ax.set_xticks(x)
    ax.set_xticklabels(df_plot.index, rotation=45, ha='right')
    ax.set_xlabel("期間")
    ylabel = "共起頻度" if not normalize else "割合"
    ax.set_ylabel(ylabel)
    ax.set_title("候補共起ペアの期間ごとの構成（積み上げ棒）")

    if normalize:
        ax.set_ylim(0.0, 1.0)
        import matplotlib.ticker as mticker
        ax.yaxis.set_major_locator(mticker.MaxNLocator(11))
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.2f'))

    # 凡例を取得して、積み上げの見た目（上→下）に対応するように順序を逆転して表示
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(handles[::-1], labels[::-1], loc='upper left', bbox_to_anchor=(1.02, 1), borderaxespad=0., fontsize=9)

    plt.tight_layout()
    if save_path is not None:
        save_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(save_path, dpi=200, bbox_inches='tight')
        print(f"積み上げ棒グラフを保存しました: {save_path}")
    plt.show()
    plt.close(fig)

# =========================================================
# 共起ネットワーク図（サブプロット -> PNG 保存対応）
# =========================================================
def plot_cooccurrence_network_subplot(ax, G: nx.Graph, period_label: str, freq: Counter, classifier: ModalityClassifier, top_n_edges: int = DRAW_TOP_EDGES):
    edges_sorted = sorted(G.edges(data=True), key=lambda x: x[2]['weight'], reverse=True)
    G_sub = nx.Graph()
    for w1, w2, data in edges_sorted[:top_n_edges]:
        G_sub.add_edge(w1, w2, weight=data['weight'])
    if len(G_sub.nodes()) == 0:
        return
    pos = nx.spring_layout(G_sub, k=0.4, iterations=50, weight='None', seed=42)
    node_sizes = [max(50, freq.get(node, 1) * 50) for node in G_sub.nodes()]
    node_colors = [TEGUCHI_COLOR if classifier.is_teguchi(node) else PROBLEM_COLOR for node in G_sub.nodes()]
    widths = [max(0.5, d['weight'] * 0.1) for _, _, d in G_sub.edges(data=True)]
    nx.draw_networkx_edges(G_sub, pos, width=widths, alpha=0.5)
    nx.draw_networkx_nodes(G_sub, pos, node_color=node_colors, node_size=node_sizes, alpha=0.25)
    nx.draw_networkx_labels(G_sub, pos, font_family=FONT_FAMILY, font_size=12)
    ax.set_title(period_label, fontsize=12)
    ax.axis("off")
    legend_handles = [
        mpatches.Patch(color=TEGUCHI_COLOR, label="手口"),
        mpatches.Patch(color=PROBLEM_COLOR, label="問題")
    ]
    ax.legend(handles=legend_handles, loc="lower left", fontsize=9, frameon=False)

def select_period_indices_for_saving(n_periods: int, max_plots: int = MAX_NETWORK_PLOTS):
    if n_periods <= max_plots:
        return list(range(n_periods))
    step = (n_periods - 1) / (max_plots - 1)
    indices = [int(round(i * step)) for i in range(max_plots)]
    uniq = sorted(set(min(n_periods-1, max(0, idx)) for idx in indices))
    i = 0
    while len(uniq) < max_plots and i < n_periods:
        if i not in uniq:
            uniq.append(i)
        i += 1
    return sorted(uniq[:max_plots])

# =========================================================
# main（実行部）
# =========================================================
def main():
    TEGUCHI_MAP = {
        "電話": ["電話","携帯"],
        "訪問": ["訪問"],
        "ネット": ["ネット", "サイト"],
        "メール": ["メール"]
    }
    classifier = ModalityClassifier(TEGUCHI_MAP)

    import warnings
    warnings.filterwarnings("ignore")
    plt.rcParams["font.family"] = FONT_FAMILY
    plt.rcParams["axes.unicode_minus"] = False

    preprocessor = TextPreprocessor()
    df = load_csv()
    periods = split_periods(df)

    # 追加表示: 期間ごとのデータ数（行数）をターミナル出力
    print("=== 期間（2か月毎）ごとのデータ数（行数） ===")
    if not periods:
        print("期間が見つかりません（データの日時列が正しくパースできているか確認してください）。")
    else:
        total_rows = 0
        for idx, (start, end, chunk) in enumerate(periods):
            count = len(chunk)
            total_rows += count
            label = f"{start.strftime('%Y-%m')}~{(start + pd.DateOffset(months=1)).strftime('%m')}"
            print(f"[{idx+1:02d}] {label} | {start.date()} 〜 {end.date()} : {count} 行")
        print(f"合計行数（期間内合計）: {total_rows} 行")
        print("========================================\n")

    period_labels = []
    period_tokenized_docs = []
    period_token_freqs = []
    period_cooccurrence_counters = []
    period_doc_counts = []  # 追加: 期間ごとの行数（相談件数）

    for start, end, chunk in periods:
        # 期間ラベル（元のまま）
        next_month = start + pd.DateOffset(months=1)
        label = f"{start.strftime('%Y-%m')}~{next_month.strftime('%m')}"
        docs = chunk["text"].dropna().tolist()

        # トークン化して共起を作る
        tokenized_docs, freq = tokenize_corpus(docs, preprocessor)
        coocc = build_cooccurrence_counter(tokenized_docs, window_size=COOCCURRENCE_WINDOW)
        period_labels.append(label)
        period_tokenized_docs.append(tokenized_docs)
        period_token_freqs.append(freq)
        period_cooccurrence_counters.append(coocc)
        period_doc_counts.append(len(chunk))  # 追加: 期間の行数（相談件数）を保存

        # デバッグ用の追加情報（必要ならコメントアウト外す）
        if DEBUG:
            print(f"{label} - raw rows: {len(chunk)}, tokenized docs: {len(tokenized_docs)}, uniq tokens: {len(freq)}")

    NETWORKS_DIR.mkdir(parents=True, exist_ok=True)
    CHARTS_DIR.mkdir(parents=True, exist_ok=True)
    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    n_periods = len(periods)
    indices_to_save = select_period_indices_for_saving(n_periods, max_plots=MAX_NETWORK_PLOTS)

    for i, (start, end, chunk) in enumerate(periods):
        if chunk.empty:
            continue
        label = period_labels[i]
        freq = period_token_freqs[i]
        G = build_cooccurrence_network_from_counter(period_cooccurrence_counters[i], min_freq=COOCCURRENCE_MIN_FREQ)
        if len(G.nodes) == 0:
            continue

        fig, ax = plt.subplots(figsize=(12, 10))
        plot_cooccurrence_network_subplot(ax, G, label, freq, classifier, top_n_edges=DRAW_TOP_EDGES)
        plt.tight_layout()

        plt.show()

        if i in indices_to_save:
            filename = NETWORKS_DIR / f"network_{i+1}_{label}.png"
            fig.savefig(filename, dpi=200, bbox_inches='tight')
            print(f"ネットワーク図を保存しました: {filename}")
        plt.close(fig)

        summaries = summarize_top_problem_sets_with_others(G, classifier, top_k=TOP_PROBLEM_SET_PER_TEGUCHI)
        print(f"\n=== {label} 手口別 上位問題セット（最大{TOP_PROBLEM_SET_PER_TEGUCHI}件） ===")
        if not summaries:
            print("（該当する手口ノードが図に存在しません）")
        for t, items in summaries.items():
            print(f"手口: {t}")
            for problem_tuple, score in items:
                label_name = "-".join(problem_tuple)
                print(f"  問題: {label_name}  (スコア合計: {score:.1f})")
        print("========================================================\n")

    # ここで上位 1-20 を一括取得して、1-10 と 11-20 を分割して出力する
    df_cooc_display_all, df_cooc_keys_all = build_cooccurrence_top_table(period_labels, period_cooccurrence_counters, top_n=TOP_COOC_PER_PERIOD*2)

    # 1-10
    df_cooc_display_top = df_cooc_display_all.iloc[:TOP_COOC_PER_PERIOD]
    df_cooc_keys_top = df_cooc_keys_all.iloc[:TOP_COOC_PER_PERIOD]
    # 11-20
    df_cooc_display_11_20 = df_cooc_display_all.iloc[TOP_COOC_PER_PERIOD:TOP_COOC_PER_PERIOD*2]
    df_cooc_keys_11_20 = df_cooc_keys_all.iloc[TOP_COOC_PER_PERIOD:TOP_COOC_PER_PERIOD*2]

    # 表 (1-10)
    table_png_path = TABLES_DIR / "table_01_10_pairs.png"
    show_cooccurrence_table(df_cooc_display_top, df_cooc_keys_top, alpha=0.25, save_path=table_png_path)
    # 表 (11-20)
    table_png_path_11_20 = TABLES_DIR / "table_11_20_pairs.png"
    show_cooccurrence_table(df_cooc_display_11_20, df_cooc_keys_11_20, alpha=0.25, save_path=table_png_path_11_20)

    # Excel 出力: 1-10 と 11-20 を別ファイルで保存
    excel_path = export_cooccurrence_table_to_excel(df_cooc_keys_top, df_cooc_display_top, filename="cooccurrence_top_1-10.xlsx", alpha=0.25, size_multiplier=1.0)
    excel_path2 = export_cooccurrence_table_to_excel(df_cooc_keys_11_20, df_cooc_display_11_20, filename="cooccurrence_11-20.xlsx", alpha=0.25, size_multiplier=1.0)

    total_counter = Counter()
    for c in period_cooccurrence_counters:
        total_counter.update(c)

    candidate_pairs = []
    table_pairs = []
    if df_cooc_keys_top is not None:
        for col in df_cooc_keys_top.columns:
            for val in df_cooc_keys_top[col].values:
                if val is None:
                    continue
                s = str(val).strip()
                if not s:
                    continue
                cleaned = _normalize_key(s).replace("｜", "|")
                table_pairs.append(cleaned)
    table_pairs = sorted({p for p in table_pairs if p})
    table_pairs_tuples = [tuple(p.split('|')) for p in table_pairs if '|' in p]

    if table_pairs_tuples:
        ordered = [p for p, _ in total_counter.most_common() if p in table_pairs_tuples]
        remaining = [p for p in table_pairs_tuples if p not in ordered]
        ordered.extend(remaining)
        for p in ordered:
            if p not in candidate_pairs:
                candidate_pairs.append(p)
            if len(candidate_pairs) >= MAX_TIMESERIES:
                break

    if len(candidate_pairs) < MAX_TIMESERIES:
        for pair, _ in total_counter.most_common():
            if pair in candidate_pairs:
                continue
            candidate_pairs.append(pair)
            if len(candidate_pairs) >= MAX_TIMESERIES:
                break

    if not candidate_pairs:
        top_overall = []
        for pair, _ in total_counter.most_common():
            top_overall.append(pair)
            if len(top_overall) >= MAX_TIMESERIES:
                break
        candidate_pairs = top_overall

    # 変更: period_doc_counts を渡す（各期間の行数を分母にして割合を計算）
    df_pairs = build_pair_timeseries_with_norm(period_labels, period_cooccurrence_counters, candidate_pairs, period_doc_counts, normalize=NORMALIZE_TIMESERIES)

    # NOTE: 以下の行は "棒グラフ（積み上げ）を一旦出力しない" 要望によりコメントアウトしました。
    # timeseries_png_path = CHARTS_DIR / "timeseries.png"
    # plot_pair_stacked_bars(df_pairs, top_n=None, normalize=NORMALIZE_TIMESERIES, save_path=timeseries_png_path)

    print("\n全ファイルはフォルダに保存されました: ", OUTPUT_DIR.resolve())

if __name__ == "__main__":
    main()
